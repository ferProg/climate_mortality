from __future__ import annotations

import argparse
import logging
from datetime import datetime
from pathlib import Path

import cdsapi
import pandas as pd
import xarray as xr
from openpyxl import Workbook, load_workbook


LOGGER = logging.getLogger("build_cams_2025_airq_excel")
EXPECTED_OBS_PER_DAY = 4  # 00, 06, 12, 18
DATASET_NAME = "cams-global-atmospheric-composition-forecasts"

ISLAND_CONFIG = {
    "tenerife": {"code": "tfe", "area": [29.0, -17.0, 27.5, -16.0]},
    "gran_canaria": {"code": "gcan", "area": [28.5, -16.2, 27.5, -15.2]},
    "lanzarote": {"code": "lzt", "area": [29.5, -14.2, 28.7, -13.1]},
    "fuerteventura": {"code": "ftv", "area": [29.0, -14.6, 27.8, -13.5]},
    "la_palma": {"code": "lpa", "area": [29.1, -18.2, 28.4, -17.7]},
    "gomera": {"code": "gom", "area": [28.3, -17.45, 27.95, -17.05]},
    "hierro": {"code": "hie", "area": [27.95, -18.25, 27.55, -17.85]},
}

OUTPUT_COLUMNS = [
    "Fecha",
    "Hora",
    "SO2",
    "NO2",
    "PM10",
    "PM2,5",
    "O3",
    "CO mg/m3",
    "Benceno",
    "Tolueno",
    "Xileno",
    "TRS",
]


def setup_logging(log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download CAMS PM10/PM2.5 for one island and write/update a 2025 Excel workbook sheet."
    )
    parser.add_argument("--island", required=True, choices=sorted(ISLAND_CONFIG.keys()))
    parser.add_argument("--start", required=True, help="Start date YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="End date YYYY-MM-DD")
    parser.add_argument(
        "--workbook",
        required=True,
        help=r'Path to output workbook, e.g. C:\data\Air_Polution_GC_2015_2025_raw\Datos2016_2025\Datos2025\Datos 2025.xlsx',
    )
    parser.add_argument(
        "--raw-dir",
        default="data/raw/cams_temp",
        help="Directory where temporary NetCDF files will be stored during download.",
    )
    parser.add_argument(
        "--sheet-name",
        default=None,
        help="Excel sheet name. Defaults to the island name.",
    )
    parser.add_argument(
        "--chunk-months",
        type=int,
        default=1,
        choices=[1, 2, 3, 6, 12],
        help="Months per CAMS request chunk.",
    )
    parser.add_argument(
        "--keep-nc",
        action="store_true",
        help="Keep downloaded NetCDF files instead of deleting them after processing.",
    )
    parser.add_argument(
        "--replace-sheet",
        action="store_true",
        help="Replace the target sheet if it already exists.",
    )
    parser.add_argument(
        "--log-dir",
        default=None,
        help="Directory for the sanity log. Defaults to the workbook folder.",
    )
    return parser.parse_args()


def make_chunks(start: pd.Timestamp, end: pd.Timestamp, chunk_months: int) -> list[tuple[pd.Timestamp, pd.Timestamp]]:
    chunks: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    current = start.normalize()
    end = end.normalize()

    while current <= end:
        next_start = (current + pd.DateOffset(months=chunk_months)).normalize()
        chunk_end = min(end, next_start - pd.Timedelta(days=1))
        chunks.append((current, chunk_end))
        current = next_start

    return chunks


def download_cams_chunk(
    client: cdsapi.Client,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    area: list[float],
    target_path: Path,
) -> Path:
    request = {
        "variable": ["particulate_matter_10um", "particulate_matter_2.5um"],
        "date": [f"{start_date.date()}/{end_date.date()}"],
        "time": ["00:00", "06:00", "12:00", "18:00"],
        "type": ["analysis"],
        "format": "netcdf",
        "area": area,
    }

    target_path.parent.mkdir(parents=True, exist_ok=True)
    LOGGER.info("Downloading CAMS %s -> %s", f"{start_date.date()}..{end_date.date()}", target_path)
    client.retrieve(DATASET_NAME, request, str(target_path))
    return target_path


def nc_to_six_hourly_df(nc_path: Path) -> pd.DataFrame:
    ds = xr.open_dataset(nc_path)

    required_vars = {"pm10", "pm2p5"}
    missing = required_vars - set(ds.data_vars)
    if missing:
        raise ValueError(f"Missing vars in {nc_path.name}: {sorted(missing)}")

    df = (
        ds[["pm10", "pm2p5"]]
        .mean(dim=["latitude", "longitude"])
        .to_dataframe()
        .reset_index()
    )

    if "valid_time" not in df.columns:
        raise ValueError(f"'valid_time' not found in {nc_path.name}")

    df["valid_time"] = pd.to_datetime(df["valid_time"])

    df = df.drop(
        columns=[c for c in ["forecast_period", "forecast_reference_time"] if c in df.columns],
        errors="ignore",
    )

    df = df.rename(columns={"valid_time": "datetime", "pm10": "PM10", "pm2p5": "PM2,5"})
    df = df[["datetime", "PM10", "PM2,5"]].copy()

    # CAMS global PM comes in kg/m3 -> convert to µg/m3
    df["PM10"] = df["PM10"] * 1_000_000_000
    df["PM2,5"] = df["PM2,5"] * 1_000_000_000

    df = df.sort_values("datetime").drop_duplicates(subset=["datetime"]).reset_index(drop=True)
    return df


def concat_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    if not frames:
        raise RuntimeError("No CAMS dataframes produced.")
    df = pd.concat(frames, ignore_index=True)
    df = df.sort_values("datetime").drop_duplicates(subset=["datetime"]).reset_index(drop=True)
    return df


def reshape_for_excel(df_6h: pd.DataFrame) -> pd.DataFrame:
    df = df_6h.copy()
    df["Fecha"] = df["datetime"].dt.strftime("%d/%m/%Y")
    df["Hora"] = df["datetime"].dt.strftime("%H:%M")

    # En este workbook solo rellenamos PM10 y PM2,5.
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = pd.NA

    df = df[OUTPUT_COLUMNS].copy()
    return df


def sanity_checks(df_6h: pd.DataFrame, island: str, start: pd.Timestamp, end: pd.Timestamp) -> None:
    expected_datetimes = pd.date_range(start.normalize(), end.normalize() + pd.Timedelta(hours=18), freq="6H")
    expected_datetimes = expected_datetimes[
        (expected_datetimes.hour.isin([0, 6, 12, 18]))
        & (expected_datetimes >= start.normalize())
        & (expected_datetimes <= end.normalize() + pd.Timedelta(hours=18))
    ]

    LOGGER.info("SANITY island=%s", island)
    LOGGER.info("Rows=%s", len(df_6h))
    LOGGER.info("Min datetime=%s", df_6h['datetime'].min())
    LOGGER.info("Max datetime=%s", df_6h['datetime'].max())
    LOGGER.info("Duplicate datetimes=%s", int(df_6h['datetime'].duplicated().sum()))
    LOGGER.info("Null datetimes=%s", int(df_6h['datetime'].isna().sum()))
    LOGGER.info("Null PM10=%s", int(df_6h['PM10'].isna().sum()))
    LOGGER.info("Null PM2,5=%s", int(df_6h['PM2,5'].isna().sum()))
    LOGGER.info("Expected 6-hour slots (naive count)=%s", len(expected_datetimes))
    LOGGER.info("Observed 6-hour slots=%s", df_6h['datetime'].nunique())
    LOGGER.info("Coverage ratio=%.4f", df_6h['datetime'].nunique() / max(len(expected_datetimes), 1))

    invalid_hours = sorted(set(df_6h["datetime"].dt.strftime("%H:%M")) - {"00:00", "06:00", "12:00", "18:00"})
    LOGGER.info("Unexpected hour labels=%s", invalid_hours)

    pm10_stats = df_6h["PM10"].describe(percentiles=[0.25, 0.5, 0.75, 0.9]).to_dict()
    pm25_stats = df_6h["PM2,5"].describe(percentiles=[0.25, 0.5, 0.75, 0.9]).to_dict()
    LOGGER.info("PM10 describe=%s", {k: round(v, 3) if pd.notna(v) else None for k, v in pm10_stats.items()})
    LOGGER.info("PM2,5 describe=%s", {k: round(v, 3) if pd.notna(v) else None for k, v in pm25_stats.items()})


def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "README"
        ws["A1"] = "Workbook created by build_cams_2025_airq_excel.py"
        ws["A2"] = "Each island sheet has row 1 = island name, row 2 = headers."
        wb.save(path)


def write_sheet(workbook_path: Path, sheet_name: str, island_label: str, df_excel: pd.DataFrame, replace_sheet: bool) -> None:
    ensure_workbook(workbook_path)
    wb = load_workbook(workbook_path)

    if sheet_name in wb.sheetnames:
        if not replace_sheet:
            raise ValueError(
                f"Sheet '{sheet_name}' already exists in {workbook_path}. Use --replace-sheet to overwrite it."
            )
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    # Row 1 mimics current files: station/island name header.
    ws["A1"] = island_label

    # Row 2: actual headers
    for idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=2, column=idx, value=col)

    # Row 3 onwards: data
    for row_idx, row in enumerate(df_excel.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=None if pd.isna(value) else value)

    # Remove default README if this workbook only contains README + target sheet
    if "README" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["README"]

    wb.save(workbook_path)
    LOGGER.info("Saved sheet '%s' -> %s", sheet_name, workbook_path)


def main() -> None:
    args = parse_args()

    start = pd.Timestamp(args.start)
    end = pd.Timestamp(args.end)

    if end < start:
        raise ValueError("--end must be >= --start")

    workbook_path = Path(args.workbook)
    log_dir = Path(args.log_dir) if args.log_dir else workbook_path.parent
    log_name = f"log_saneamiento_ingest_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{args.island}.txt"
    setup_logging(log_dir / log_name)

    sheet_name = args.sheet_name or args.island
    island_cfg = ISLAND_CONFIG[args.island]
    code = island_cfg["code"]
    area = island_cfg["area"]
    raw_dir = Path(args.raw_dir)

    LOGGER.info("START island=%s code=%s start=%s end=%s", args.island, code, start.date(), end.date())
    LOGGER.info("Workbook=%s", workbook_path)
    LOGGER.info("Sheet=%s", sheet_name)

    client = cdsapi.Client()
    chunks = make_chunks(start, end, args.chunk_months)
    LOGGER.info("Chunks=%s", len(chunks))

    nc_paths: list[Path] = []
    frames: list[pd.DataFrame] = []

    for chunk_start, chunk_end in chunks:
        nc_name = f"cams_pm_{code}_{chunk_start.date()}_{chunk_end.date()}.nc"
        nc_path = raw_dir / args.island / nc_name
        download_cams_chunk(client, chunk_start, chunk_end, area, nc_path)
        nc_paths.append(nc_path)
        frames.append(nc_to_six_hourly_df(nc_path))

    df_6h = concat_frames(frames)
    sanity_checks(df_6h, args.island, start, end)
    df_excel = reshape_for_excel(df_6h)
    write_sheet(
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        island_label=args.island,
        df_excel=df_excel,
        replace_sheet=args.replace_sheet,
    )

    if not args.keep_nc:
        for path in nc_paths:
            try:
                path.unlink(missing_ok=True)
                LOGGER.info("Deleted temp file=%s", path)
            except Exception as exc:
                LOGGER.warning("Could not delete %s: %s", path, exc)

    LOGGER.info("DONE island=%s rows_written=%s", args.island, len(df_excel))


if __name__ == "__main__":
    main()
